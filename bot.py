import telebot
from telebot import types
import requests
from datetime import datetime, timedelta
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WEBHOOK = 'https://altuspro.bitrix24.ru/rest/33/1r23rn74ww8yq892/'
BOT_TOKEN = '8612831715:AAE1OVngy867YStfZhEZUiyqcqbEAt_8ZA0'

bot = telebot.TeleBot(BOT_TOKEN)

WON_STAGES = ['WON', 'UC_ZWS97R', 'EXECUTING']
LOST_STAGES = ['LOSE', 'APOLOGY']
MISSED_STAGES = ['NEW', '1']
EXPECTED_STAGES = ['PREPAYMENT_INVOICE']

STAGE_NAMES = {
    'NEW': 'Потенциальная потребность',
    '1': 'Потребность подтверждена',
    'UC_FZ1R5C': 'ДЛЯ АНИ!!',
    'PREPARATION': 'ТКП направлено',
    '3': 'ТКП перекупам',
    '2': 'Работа с возражениями',
    'PREPAYMENT_INVOICE': 'Счет на предоплату',
    'EXECUTING': 'В работе - счет оплачен',
    'UC_ZWS97R': 'Отгружен БЕЗ ДОКУМЕНТОВ',
    'UC_018IHX': 'ПРОРАБОТАТЬ',
    '6': 'Напомнить',
    '5': 'нет бюджета',
    'UC_IOD5R7': 'Документооборот',
    'UC_H0H6EG': 'Комплексная закупка',
    '4': 'Тендер',
    'UC_WLPIEC': 'Заказ до 10к/Озон',
    'UC_3W24DF': 'Не рассматривают аналоги',
    'WON': 'Сделка успешна',
    'LOSE': 'Сделка провалена',
    'APOLOGY': 'Спам'
}

user_state = {}

def main_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        types.KeyboardButton('📊 Все менеджеры'),
        types.KeyboardButton('👤 По менеджеру'),
        types.KeyboardButton('💰 Выручка'),
        types.KeyboardButton('⏳ Ожидаем оплату'),
        types.KeyboardButton('🔴 Пропущенные'),
        types.KeyboardButton('⚠️ Требуют внимания'),
        types.KeyboardButton('📋 Excel отчёт')
    )
    return kb

def period_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        types.KeyboardButton('📅 Сегодня'),
        types.KeyboardButton('📅 Неделя'),
        types.KeyboardButton('📅 Месяц'),
        types.KeyboardButton('📅 Квартал'),
        types.KeyboardButton('◀️ Назад')
    )
    return kb

def bx(method, params=None):
    url = WEBHOOK + method + '.json'
    r = requests.get(url, params=params or {})
    return r.json().get('result', [])

def get_users():
    users = bx('user.get', {'select[]': ['ID', 'NAME', 'LAST_NAME']})
    return {u['ID']: (u.get('NAME', '') + ' ' + u.get('LAST_NAME', '')).strip() for u in users}

def get_deals_for_month():
    """Загружает сделки только за текущий месяц"""
    deals = []
    start = 0
    users = get_users()
    now = datetime.now()
    date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    date_filter = date_from.strftime('%d.%m.%Y')
    while True:
        params = {
            'select[]': ['ID', 'TITLE', 'STAGE_ID', 'ASSIGNED_BY_ID', 'OPPORTUNITY', 'DATE_CREATE', 'CLOSEDATE', 'SOURCE_ID'],
            'filter[>=DATE_CREATE]': date_filter,
            'order[DATE_CREATE]': 'DESC',
            'start': start
        }
        result = bx('crm.deal.list', params)
        if not result:
            break
        for d in result:
            d['MANAGER'] = users.get(d.get('ASSIGNED_BY_ID', ''), 'Не назначен')
        deals.extend(result)
        if len(result) < 50:
            break
        start += 50
        if len(deals) > 2000:
            break
    return deals

def get_deals():
    """Загружает сделки за 3 месяца"""
    deals = []
    start = 0
    users = get_users()
    date_from = datetime.now() - timedelta(days=90)
    date_filter = date_from.strftime('%d.%m.%Y')
    while True:
        params = {
            'select[]': ['ID', 'TITLE', 'STAGE_ID', 'ASSIGNED_BY_ID', 'OPPORTUNITY', 'DATE_CREATE', 'CLOSEDATE', 'SOURCE_ID'],
            'filter[>=DATE_CREATE]': date_filter,
            'order[DATE_CREATE]': 'DESC',
            'start': start
        }
        result = bx('crm.deal.list', params)
        if not result:
            break
        for d in result:
            d['MANAGER'] = users.get(d.get('ASSIGNED_BY_ID', ''), 'Не назначен')
        deals.extend(result)
        if len(result) < 50:
            break
        start += 50
        if len(deals) > 5000:
            break
    return deals

def get_period(key):
    now = datetime.now()
    if key == '📅 Сегодня':
        return now.replace(hour=0, minute=0, second=0, microsecond=0), now
    if key == '📅 Неделя':
        return now - timedelta(days=7), now
    if key == '📅 Месяц':
        return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0), now
    if key == '📅 Квартал':
        return now - timedelta(days=90), now
    return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0), now

def get_deal_date(d):
    date_str = d.get('DATE_CREATE', '')
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str.replace('+03:00', '').replace('T', ' ').split('.')[0])
    except:
        return None

def filter_by_period(deals, date_from, date_to):
    if not date_from:
        return deals
    result = []
    for d in deals:
        dt = get_deal_date(d)
        if dt and date_from <= dt <= date_to:
            result.append(d)
    return result

def days_since(date_str):
    if not date_str:
        return 999
    try:
        d = datetime.fromisoformat(date_str.replace('+03:00', '').replace('T', ' ').split('.')[0])
        return (datetime.now() - d).days
    except:
        return 999

def format_money(v):
    try:
        n = float(v or 0)
        if n > 0:
            return f"{int(n):,}".replace(',', ' ') + ' ₽'
    except:
        pass
    return '—'

def format_period(date_from, date_to):
    if not date_from:
        return 'все время'
    return f"{date_from.strftime('%d.%m')}–{date_to.strftime('%d.%m.%Y')}"

def sum_revenue(deals_list):
    return sum(float(d.get('OPPORTUNITY', 0) or 0) for d in deals_list)

def get_managers_list(deals):
    managers = set()
    for d in deals:
        m = d.get('MANAGER', '')
        if m and m != 'Не назначен':
            managers.add(m)
    return sorted(managers)

def managers_menu(deals):
    managers = get_managers_list(deals)
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    for m in managers:
        kb.add(types.KeyboardButton(f'👤 {m}'))
    kb.add(types.KeyboardButton('◀️ Назад'))
    return kb, managers

def report_managers_menu(deals):
    managers = get_managers_list(deals)
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    kb.add(types.KeyboardButton('📊 Общий отчёт (все менеджеры)'))
    for m in managers:
        kb.add(types.KeyboardButton(f'📋 {m}'))
    kb.add(types.KeyboardButton('◀️ Назад'))
    return kb

def show_all(chat_id, deals, date_from, date_to):
    filtered = filter_by_period(deals, date_from, date_to)
    managers = {}
    for d in filtered:
        m = d.get('MANAGER', 'Не назначен')
        if m not in managers:
            managers[m] = {'won_f': 0, 'won_s': 0, 'won_p': 0, 'active': 0, 'lost': 0, 'expected': 0, 'revenue': 0, 'expected_sum': 0}
        sid = d.get('STAGE_ID')
        opp = float(d.get('OPPORTUNITY', 0) or 0)
        if sid == 'WON':
            managers[m]['won_f'] += 1; managers[m]['revenue'] += opp
        elif sid == 'UC_ZWS97R':
            managers[m]['won_s'] += 1; managers[m]['revenue'] += opp
        elif sid == 'EXECUTING':
            managers[m]['won_p'] += 1; managers[m]['revenue'] += opp
        elif sid in LOST_STAGES:
            managers[m]['lost'] += 1
        elif sid in EXPECTED_STAGES:
            managers[m]['expected'] += 1; managers[m]['expected_sum'] += opp; managers[m]['active'] += 1
        else:
            managers[m]['active'] += 1

    period_str = format_period(date_from, date_to)
    text = f"📊 *Все менеджеры* | {period_str}\nВсего: {len(filtered)} сделок\n\n"

    for name, s in sorted(managers.items(), key=lambda x: x[1]['revenue'], reverse=True):
        won_total = s['won_f'] + s['won_s'] + s['won_p']
        total = won_total + s['active'] + s['lost']
        conv = round(won_total / max(total, 1) * 100)
        text += f"👤 *{name}*\n"
        text += f"  ✅{s['won_f']} 📦{s['won_s']} 💳{s['won_p']} | 💰{format_money(s['revenue'])}\n"
        if s['expected'] > 0:
            text += f"  ⏳ Ожидаем: {s['expected']} ({format_money(s['expected_sum'])})\n"
        text += f"  🔄{s['active']} ❌{s['lost']} 📈{conv}%\n\n"

    bot.send_message(chat_id, text, parse_mode='Markdown', reply_markup=main_menu())

def show_manager(chat_id, manager_name, deals, date_from, date_to):
    filtered = [d for d in deals if manager_name in d.get('MANAGER', '')]
    if date_from:
        filtered = filter_by_period(filtered, date_from, date_to)

    won_final = [d for d in filtered if d.get('STAGE_ID') == 'WON']
    won_shipped = [d for d in filtered if d.get('STAGE_ID') == 'UC_ZWS97R']
    won_paid = [d for d in filtered if d.get('STAGE_ID') == 'EXECUTING']
    won = won_final + won_shipped + won_paid
    lost = [d for d in filtered if d.get('STAGE_ID') in LOST_STAGES]
    expected = [d for d in filtered if d.get('STAGE_ID') in EXPECTED_STAGES]
    active = [d for d in filtered if d.get('STAGE_ID') not in WON_STAGES + LOST_STAGES]
    missed = [d for d in active if d.get('STAGE_ID') in MISSED_STAGES and days_since(d.get('DATE_CREATE', '')) >= 3]

    total_revenue = sum_revenue(won)
    expected_sum = sum_revenue(expected)
    conv = round(len(won) / max(len(filtered), 1) * 100)

    stages = {}
    for d in active:
        s = STAGE_NAMES.get(d.get('STAGE_ID', ''), d.get('STAGE_ID', '—'))
        stages[s] = stages.get(s, 0) + 1

    full_name = filtered[0].get('MANAGER', manager_name) if filtered else manager_name
    period_str = format_period(date_from, date_to)

    text = f"👤 *{full_name}*\n📅 {period_str}\n\n"
    text += f"📊 Всего сделок: {len(filtered)}\n\n"
    text += f"💰 *Оплаты:*\n"
    text += f"  ✅ Сделка успешна: {len(won_final)} ({format_money(sum_revenue(won_final))})\n"
    text += f"  📦 Отгружен без документов: {len(won_shipped)} ({format_money(sum_revenue(won_shipped))})\n"
    text += f"  💳 Счет оплачен: {len(won_paid)} ({format_money(sum_revenue(won_paid))})\n"
    text += f"  📈 Итого: {len(won)} ({format_money(total_revenue)})\n\n"

    if expected:
        text += f"⏳ Ожидаем оплату: {len(expected)} ({format_money(expected_sum)})\n\n"

    text += f"❌ Провалов: {len(lost)}\n"
    text += f"🔄 В работе: {len(active)}\n"
    text += f"📈 Конверсия: {conv}%\n"

    if stages:
        text += f"\n📋 *По стадиям:*\n"
        for stage, count in sorted(stages.items(), key=lambda x: x[1], reverse=True)[:8]:
            text += f"  • {stage}: {count}\n"

    if missed:
        text += f"\n🔴 *Пропущенных: {len(missed)}*\n"
        for d in missed[:5]:
            days = days_since(d.get('DATE_CREATE', ''))
            text += f"  • {d.get('TITLE', '—')} ({days} дн.)\n"

    bot.send_message(chat_id, text, parse_mode='Markdown', reply_markup=main_menu())

def show_revenue(chat_id, deals, date_from, date_to):
    filtered = filter_by_period(deals, date_from, date_to)
    won = [d for d in filtered if d.get('STAGE_ID') in WON_STAGES]
    total = sum_revenue(won)
    by_manager = {}
    for d in won:
        m = d.get('MANAGER', '—')
        by_manager[m] = by_manager.get(m, 0) + float(d.get('OPPORTUNITY', 0) or 0)

    period_str = format_period(date_from, date_to)
    text = f"💰 *Выручка* | {period_str}\nИтого: *{format_money(total)}*\n\n"
    for name, rev in sorted(by_manager.items(), key=lambda x: x[1], reverse=True):
        text += f"👤 {name}: {format_money(rev)}\n"

    bot.send_message(chat_id, text, parse_mode='Markdown', reply_markup=main_menu())

def show_expected(chat_id, deals, date_from, date_to):
    filtered = filter_by_period(deals, date_from, date_to)
    expected = [d for d in filtered if d.get('STAGE_ID') in EXPECTED_STAGES]

    if not expected:
        bot.send_message(chat_id, "✅ Нет сделок в ожидании оплаты!", reply_markup=main_menu())
        return

    total = sum_revenue(expected)
    period_str = format_period(date_from, date_to)
    by_manager = {}
    for d in expected:
        m = d.get('MANAGER', '—')
        if m not in by_manager:
            by_manager[m] = []
        by_manager[m].append(d)

    text = f"⏳ *Ожидаемые оплаты* | {period_str}\nИтого: *{format_money(total)}* ({len(expected)} сделок)\n\n"
    for name, mgr_deals in sorted(by_manager.items()):
        mgr_sum = sum_revenue(mgr_deals)
        text += f"👤 *{name}* — {format_money(mgr_sum)}\n"
        for d in mgr_deals[:5]:
            text += f"  • {d.get('TITLE', '—')} — {format_money(d.get('OPPORTUNITY', 0))}\n"
        text += "\n"

    bot.send_message(chat_id, text, parse_mode='Markdown', reply_markup=main_menu())

def show_missed(chat_id, deals):
    missed = [d for d in deals if d.get('STAGE_ID') in MISSED_STAGES and days_since(d.get('DATE_CREATE', '')) >= 3]

    if not missed:
        bot.send_message(chat_id, "🎉 Пропущенных заявок нет!", reply_markup=main_menu())
        return

    text = f"🔴 *Пропущенные ({len(missed)})*\n\n"
    for d in missed[:20]:
        days = days_since(d.get('DATE_CREATE', ''))
        emoji = "🔴" if days >= 7 else "🟡"
        text += f"{emoji} {d.get('TITLE', '—')}\n  👤 {d.get('MANAGER', '—')} | {days} дн.\n\n"

    bot.send_message(chat_id, text, parse_mode='Markdown', reply_markup=main_menu())

def show_attention(chat_id, deals):
    attention = [d for d in deals if d.get('STAGE_ID') not in WON_STAGES + LOST_STAGES and days_since(d.get('DATE_CREATE', '')) >= 7]

    if not attention:
        bot.send_message(chat_id, "✅ Все сделки в норме!", reply_markup=main_menu())
        return

    text = f"⚠️ *Требуют внимания ({len(attention)})*\n_(висят 7+ дней)_\n\n"
    for d in attention[:15]:
        days = days_since(d.get('DATE_CREATE', ''))
        stage = STAGE_NAMES.get(d.get('STAGE_ID', ''), d.get('STAGE_ID', ''))
        text += f"• {d.get('TITLE', '—')}\n  👤 {d.get('MANAGER', '—')} | {stage} | {days} дн.\n\n"

    bot.send_message(chat_id, text, parse_mode='Markdown', reply_markup=main_menu())

def create_manager_excel(deals, manager_name, date_from, date_to):
    wb = openpyxl.Workbook()
    BLUE = "1E3A5F"
    LIGHT_BLUE = "2E86AB"
    GREEN_LIGHT = "E8F5E9"
    RED_LIGHT = "FFEBEE"
    YELLOW = "FFF9C4"
    LIGHT_GRAY = "ECF0F1"
    WHITE = "FFFFFF"

    def sc(cell, bold=False, color=None, bg=None, size=11, align='left', wrap=False):
        cell.font = Font(bold=bold, size=size, color=color or "000000", name='Calibri')
        if bg:
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)

    def bc(cell):
        thin = Side(style='thin', color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Фильтруем по менеджеру и периоду
    if manager_name == 'ALL':
        filtered = filter_by_period(deals, date_from, date_to)
        title_name = 'Все менеджеры'
    else:
        filtered = [d for d in deals if manager_name in d.get('MANAGER', '')]
        filtered = filter_by_period(filtered, date_from, date_to)
        title_name = manager_name

    # Лист 1 - Сводка
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:H1')
    ws['A1'] = f'ОТЧЕТ AltusPro | {title_name}'
    sc(ws['A1'], bold=True, color=WHITE, bg=BLUE, size=14, align='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:H2')
    ws['A2'] = f'{date_from.strftime("%d.%m.%Y")} – {date_to.strftime("%d.%m.%Y")} | Сформирован: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    sc(ws['A2'], color="666666", bg=LIGHT_GRAY, size=10, align='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    # Сводка по менеджеру
    won_final = [d for d in filtered if d.get('STAGE_ID') == 'WON']
    won_shipped = [d for d in filtered if d.get('STAGE_ID') == 'UC_ZWS97R']
    won_paid = [d for d in filtered if d.get('STAGE_ID') == 'EXECUTING']
    won = won_final + won_shipped + won_paid
    lost = [d for d in filtered if d.get('STAGE_ID') in LOST_STAGES]
    expected = [d for d in filtered if d.get('STAGE_ID') in EXPECTED_STAGES]
    active = [d for d in filtered if d.get('STAGE_ID') not in WON_STAGES + LOST_STAGES]

    rev_f = sum_revenue(won_final)
    rev_s = sum_revenue(won_shipped)
    rev_p = sum_revenue(won_paid)
    rev_total = rev_f + rev_s + rev_p
    rev_e = sum_revenue(expected)
    conv = round(len(won) / max(len(filtered), 1) * 100)

    summary = [
        ('Всего сделок', len(filtered), None),
        ('', '', None),
        ('✅ Сделка успешна (шт)', len(won_final), None),
        ('✅ Сделка успешна (₽)', rev_f, '#,##0 "₽"'),
        ('📦 Отгружен без документов (шт)', len(won_shipped), None),
        ('📦 Отгружен без документов (₽)', rev_s, '#,##0 "₽"'),
        ('💳 Счет оплачен (шт)', len(won_paid), None),
        ('💳 Счет оплачен (₽)', rev_p, '#,##0 "₽"'),
        ('💰 ИТОГО ОПЛАТ (шт)', len(won), None),
        ('💰 ИТОГО ОПЛАТ (₽)', rev_total, '#,##0 "₽"'),
        ('', '', None),
        ('⏳ Ожидаем оплату (шт)', len(expected), None),
        ('⏳ Ожидаем оплату (₽)', rev_e, '#,##0 "₽"'),
        ('', '', None),
        ('🔄 В работе', len(active), None),
        ('❌ Провалов', len(lost), None),
        ('📈 Конверсия', f'{conv}%', None),
    ]

    row = 4
    for label, value, fmt in summary:
        if label == '':
            row += 1
            continue
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_v = ws.cell(row=row, column=2, value=value)
        is_total = 'ИТОГО' in label
        bg = GREEN_LIGHT if 'ИТОГО' in label else (YELLOW if 'Ожидаем' in label else WHITE)
        sc(cell_l, bold=is_total, bg=bg)
        sc(cell_v, bold=is_total, bg=bg, align='right')
        if fmt:
            cell_v.number_format = fmt
        bc(cell_l); bc(cell_v)
        ws.row_dimensions[row].height = 22
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    # Лист 2 - Детализация сделок
    ws2 = wb.create_sheet("Все сделки")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:G1')
    ws2['A1'] = f'ДЕТАЛИЗАЦИЯ | {title_name} | {date_from.strftime("%d.%m.%Y")} – {date_to.strftime("%d.%m.%Y")}'
    sc(ws2['A1'], bold=True, color=WHITE, bg=BLUE, size=13, align='center')
    ws2.row_dimensions[1].height = 30

    headers = ['Менеджер', 'Название сделки', 'Статус', 'Сумма', 'Дата создания', 'Дней в работе', 'Источник']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        sc(cell, bold=True, color=WHITE, bg=LIGHT_BLUE, size=10, align='center')
        bc(cell)
    ws2.row_dimensions[2].height = 25

    row2 = 3
    for d in sorted(filtered, key=lambda x: x.get('STAGE_ID', '')):
        sid = d.get('STAGE_ID', '')
        opp = float(d.get('OPPORTUNITY', 0) or 0)
        days = days_since(d.get('DATE_CREATE', ''))
        if sid == 'WON':
            bg = GREEN_LIGHT
        elif sid in ['UC_ZWS97R', 'EXECUTING']:
            bg = "#E3F2FD"
        elif sid in LOST_STAGES:
            bg = RED_LIGHT
        elif sid in EXPECTED_STAGES:
            bg = YELLOW
        elif days >= 7:
            bg = "#FFF3E0"
        else:
            bg = WHITE if row2 % 2 == 0 else LIGHT_GRAY

        vals = [d.get('MANAGER', '—'), d.get('TITLE', '—'), STAGE_NAMES.get(sid, sid),
                opp if opp > 0 else None,
                d.get('DATE_CREATE', '')[:10] if d.get('DATE_CREATE') else '—',
                days if days < 999 else '—',
                d.get('SOURCE_ID', '—') or '—']

        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row2, column=col, value=val)
            sc(cell, bg=bg, wrap=(col == 2))
            if col == 4 and val:
                cell.number_format = '#,##0 "₽"'
                sc(cell, bg=bg, align='right')
            bc(cell)
        row2 += 1

    for i, w in enumerate([22, 40, 22, 14, 13, 13, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Лист 3 - Ожидаемые оплаты
    ws3 = wb.create_sheet("Ожидаемые оплаты")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:E1')
    ws3['A1'] = f'ОЖИДАЕМЫЕ ОПЛАТЫ | {title_name}'
    sc(ws3['A1'], bold=True, color=WHITE, bg="E67E22", size=13, align='center')
    ws3.row_dimensions[1].height = 30

    for col, h in enumerate(['Менеджер', 'Название сделки', 'Сумма', 'Дата создания', 'Дней ожидания'], 1):
        cell = ws3.cell(row=2, column=col, value=h)
        sc(cell, bold=True, color=WHITE, bg=LIGHT_BLUE, size=10, align='center')
        bc(cell)

    r3 = 3
    total_exp = 0
    for d in sorted(expected, key=lambda x: x.get('MANAGER', '')):
        opp = float(d.get('OPPORTUNITY', 0) or 0)
        total_exp += opp
        days = days_since(d.get('DATE_CREATE', ''))
        bg = "#FFF3E0" if days > 7 else YELLOW
        for col, val in enumerate([d.get('MANAGER', '—'), d.get('TITLE', '—'),
                                    opp if opp > 0 else None,
                                    d.get('DATE_CREATE', '')[:10] if d.get('DATE_CREATE') else '—',
                                    days if days < 999 else '—'], 1):
            cell = ws3.cell(row=r3, column=col, value=val)
            sc(cell, bg=bg)
            if col == 3 and val:
                cell.number_format = '#,##0 "₽"'
                sc(cell, bg=bg, align='right')
            bc(cell)
        r3 += 1

    if expected:
        tc = ws3.cell(row=r3, column=1, value='ИТОГО')
        sc(tc, bold=True)
        tcc = ws3.cell(row=r3, column=3, value=total_exp)
        sc(tcc, bold=True, color=WHITE, bg="E67E22", align='right')
        tcc.number_format = '#,##0 "₽"'

    for i, w in enumerate([22, 40, 16, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@bot.message_handler(commands=['start'])
def handle_start(message):
    bot.send_message(message.chat.id,
                     "👋 Привет! Я CRM бот *AltusPro*\n\nВыбери что хочешь посмотреть 👇",
                     parse_mode='Markdown', reply_markup=main_menu())

@bot.message_handler(func=lambda m: True)
def handle_text(message):
    text = message.text.strip()
    chat_id = message.chat.id
    uid = message.from_user.id

    if text == '◀️ Назад':
        user_state.pop(uid, None)
        bot.send_message(chat_id, "Главное меню 👇", reply_markup=main_menu())
        return

    if uid in user_state:
        state = user_state[uid]
        action = state.get('action')

        if text in ['📅 Сегодня', '📅 Неделя', '📅 Месяц', '📅 Квартал']:
            date_from, date_to = get_period(text)
            bot.send_message(chat_id, "⏳ Загружаю данные из Битрикс24...")
            deals = get_deals()

            if action == 'all':
                show_all(chat_id, deals, date_from, date_to)
            elif action == 'revenue':
                show_revenue(chat_id, deals, date_from, date_to)
            elif action == 'expected':
                show_expected(chat_id, deals, date_from, date_to)
            elif action == 'manager' and 'manager_name' in state:
                show_manager(chat_id, state['manager_name'], deals, date_from, date_to)

            user_state.pop(uid, None)
            return

        if action == 'choose_manager':
            if text.startswith('👤 '):
                manager_name = text[2:].strip()
                user_state[uid] = {'action': 'manager', 'manager_name': manager_name}
                bot.send_message(chat_id, f"👤 *{manager_name}*\nВыбери период 👇",
                                 parse_mode='Markdown', reply_markup=period_menu())
                return

        if action == 'choose_report_manager':
            now = datetime.now()
            date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = now

            if text == '📊 Общий отчёт (все менеджеры)':
                bot.send_message(chat_id, f"⏳ Формирую общий отчёт за {date_from.strftime('%d.%m')}–{date_to.strftime('%d.%m.%Y')}...")
                deals = get_deals_for_month()
                excel = create_manager_excel(deals, 'ALL', date_from, date_to)
                filename = f"AltusPro_все_{date_from.strftime('%d.%m')}-{date_to.strftime('%d.%m.%Y')}.xlsx"
                bot.send_document(chat_id, excel, visible_file_name=filename,
                                  caption=f"📊 Общий отчёт | {date_from.strftime('%d.%m')}–{date_to.strftime('%d.%m.%Y')}")
                user_state.pop(uid, None)
                bot.send_message(chat_id, "Главное меню 👇", reply_markup=main_menu())
                return

            if text.startswith('📋 '):
                manager_name = text[2:].strip()
                bot.send_message(chat_id, f"⏳ Формирую отчёт по {manager_name}...")
                deals = get_deals_for_month()
                excel = create_manager_excel(deals, manager_name, date_from, date_to)
                safe_name = manager_name.replace(' ', '_')
                filename = f"AltusPro_{safe_name}_{date_from.strftime('%d.%m')}-{date_to.strftime('%d.%m.%Y')}.xlsx"
                bot.send_document(chat_id, excel, visible_file_name=filename,
                                  caption=f"📋 {manager_name} | {date_from.strftime('%d.%m')}–{date_to.strftime('%d.%m.%Y')}")
                user_state.pop(uid, None)
                bot.send_message(chat_id, "Главное меню 👇", reply_markup=main_menu())
                return

    if text == '📊 Все менеджеры':
        user_state[uid] = {'action': 'all'}
        bot.send_message(chat_id, "📊 Все менеджеры\nВыбери период 👇", reply_markup=period_menu())

    elif text == '👤 По менеджеру':
        bot.send_message(chat_id, "⏳ Загружаю список менеджеров...")
        deals = get_deals()
        kb, managers = managers_menu(deals)
        user_state[uid] = {'action': 'choose_manager'}
        bot.send_message(chat_id, "👤 Выбери менеджера 👇", reply_markup=kb)

    elif text == '💰 Выручка':
        user_state[uid] = {'action': 'revenue'}
        bot.send_message(chat_id, "💰 Выручка\nВыбери период 👇", reply_markup=period_menu())

    elif text == '⏳ Ожидаем оплату':
        user_state[uid] = {'action': 'expected'}
        bot.send_message(chat_id, "⏳ Ожидаемые оплаты\nВыбери период 👇", reply_markup=period_menu())

    elif text == '🔴 Пропущенные':
        bot.send_message(chat_id, "⏳ Загружаю данные из Битрикс24...")
        deals = get_deals()
        show_missed(chat_id, deals)

    elif text == '⚠️ Требуют внимания':
        bot.send_message(chat_id, "⏳ Загружаю данные из Битрикс24...")
        deals = get_deals()
        show_attention(chat_id, deals)

    elif text == '📋 Excel отчёт':
        bot.send_message(chat_id, "⏳ Загружаю список менеджеров...")
        deals = get_deals_for_month()
        kb = report_managers_menu(deals)
        user_state[uid] = {'action': 'choose_report_manager'}
        now = datetime.now()
        bot.send_message(chat_id,
                         f"📋 *Excel отчёт* за {now.replace(day=1).strftime('%d.%m')}–{now.strftime('%d.%m.%Y')}\n\nВыбери менеджера 👇",
                         parse_mode='Markdown', reply_markup=kb)

    else:
        bot.send_message(chat_id, "Используй кнопки меню 👇", reply_markup=main_menu())


if __name__ == '__main__':
    print("Бот запущен!")
    bot.infinity_polling()
